import pandas as pd
from openpyxl import load_workbook
import json
import sys

class ExcelAnalyzer:
    def __init__(self, filepath):
        self.filepath = filepath
        self.results = {"sheets": {}}

    def analyze(self):
        try:
            # Get sheet names using openpyxl (handles encoding better)
            wb = load_workbook(self.filepath, read_only=True, data_only=False)
            sheets = wb.sheetnames
            print(f"Found {len(sheets)} sheets")
            self.results["sheet_names"] = sheets

            # Analyze each sheet with pandas
            for sheet in sheets:
                print(f"Reading sheet: {sheet}")
                try:
                    df = pd.read_excel(self.filepath, sheet_name=sheet)

                    # Get column info
                    columns = list(df.columns)
                    dtypes = df.dtypes.astype(str).to_dict()

                    # Detect data types in first few rows
                    sample_data = df.head(3).to_dict(orient='records')

                    # Look for specific column types
                    has_stock_code = any('code' in str(col).lower() or 'nse' in str(col).lower() or 'symbol' in str(col).lower() for col in columns)
                    has_price = any('price' in str(col).lower() or 'cmp' in str(col).lower() or 'close' in str(col).lower() for col in columns)
                    has_dma = any('dma' in str(col).lower() for col in columns)
                    has_date = any('date' in str(col).lower() for col in columns)
                    has_pnl = any(x in str(col).lower() for col in columns for x in ['profit', 'pnl', 'gain', 'return', 'p&l'])

                    sheet_info = {
                        "rows": int(df.shape[0]),
                        "cols": int(df.shape[1]),
                        "columns": columns,
                        "dtypes": dtypes,
                        "sample": sample_data,
                        "detected_types": {
                            "likely_stock_codes": has_stock_code,
                            "has_price": has_price,
                            "has_dma": has_dma,
                            "has_date": has_date,
                            "has_pnl": has_pnl
                        }
                    }

                    # Try to get formulas (read-only)
                    ws = wb[sheet]
                    formulas = []
                    for row_idx, row in enumerate(ws.iter_rows(max_row=min(50, ws.max_row)), 1):
                        for cell in row:
                            if cell.data_type == 'f' and cell.value:
                                formulas.append({
                                    "cell": cell.coordinate,
                                    "row": row_idx,
                                    "col": cell.column_letter,
                                    "formula": str(cell.value)
                                })
                                if len(formulas) < 10:  # Only first 10 formulas
                                    pass  # Keep all we collect

                    if formulas:
                        sheet_info["formulas_sample"] = formulas[:10]

                    # Save sheet to CSV
                    csv_name = f"{sheet.replace(' ', '_').replace('/', '_').replace('\\', '_')}.csv"
                    df.to_csv(csv_name, index=False, encoding='utf-8')
                    sheet_info["csv_file"] = csv_name

                    self.results["sheets"][sheet] = sheet_info
                    print(f"  -> {df.shape[0]} rows, {df.shape[1]} cols")

                except Exception as e:
                    print(f"  Error reading sheet {sheet}: {e}")
                    self.results["sheets"][sheet] = {"error": str(e)}

            wb.close()

        except Exception as e:
            print(f"Fatal error: {e}")
            import traceback
            self.results["fatal_error"] = str(e)
            self.results["traceback"] = traceback.format_exc()

        # Save results to JSON
        with open("excel_analysis_results.json", "w", encoding="utf-8") as f:
            json.dump(self.results, f, indent=2, default=str, ensure_ascii=False)

        print("\nResults saved to: excel_analysis_results.json")
        return self.results

if __name__ == "__main__":
    filepath = r"C:\Users\bhara\OneDrive\Desktop\Trading\DMA_CAR_Enhanced_500.xlsx"
    analyzer = ExcelAnalyzer(filepath)
    analyzer.analyze()

    # Also create a summary
    print("\n" + "="*80)
    print("SUMMARY")
    print("="*80)
    print(f"Total sheets: {len(analyzer.results.get('sheets', {}))}")
    for sheet_name, info in analyzer.results.get("sheets", {}).items():
        if "error" not in info:
            print(f"  {sheet_name}: {info['rows']} rows, {info['cols']} cols")
            if info.get('detected_types'):
                types = info['detected_types']
                print(f"    Stock codes: {types['likely_stock_codes']}, Price: {types['has_price']}, DMA: {types['has_dma']}")
