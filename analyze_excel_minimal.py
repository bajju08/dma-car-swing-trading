import pandas as pd
from openpyxl import load_workbook
import json

filepath = r"C:\Users\bhara\OneDrive\Desktop\Trading\DMA_CAR_Enhanced_500.xlsx"
results = {"sheets": {}}

try:
    wb = load_workbook(filepath, read_only=True, data_only=False)
    sheets = wb.sheetnames
    print(f"Found {len(sheets)} sheets")
    results["sheet_names"] = sheets

    for sheet in sheets:
        print(f"Processing: {sheet}")
        try:
            df = pd.read_excel(filepath, sheet_name=sheet)

            columns = list(df.columns)
            sheet_info = {
                "rows": int(df.shape[0]),
                "cols": int(df.shape[1]),
                "columns": columns,
            }

            # Save to CSV
            csv_name = f"{sheet}_data.csv"
            df.to_csv(csv_name, index=False, encoding='utf-8')
            sheet_info["csv_file"] = csv_name

            results["sheets"][sheet] = sheet_info
            print(f"  -> {df.shape[0]} rows, {df.shape[1]} cols, columns: {len(columns)}")

        except Exception as e:
            print(f"  Error: {e}")
            results["sheets"][sheet] = {"error": str(e)}

    wb.close()

except Exception as e:
    print(f"Fatal: {e}")
    results["fatal_error"] = str(e)

with open("excel_analysis_results.json", "w", encoding="utf-8") as f:
    json.dump(results, f, indent=2, default=str)

print("Analysis complete. Saved to excel_analysis_results.json")
